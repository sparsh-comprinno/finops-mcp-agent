"""Excel report generation with styled sheets."""

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .core import LOG_DIR, Finding

# Styles
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_TITLE_FONT = Font(bold=True, size=14, color="1F3864")
_MONEY_FMT = '$#,##0.00'
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_PRIORITY_FILLS = {
    "HIGH": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "LOW": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}
_RISK_MAP = {
    "Database": "Medium (Requires App Validation)",
    "Networking": "Medium (Connectivity Impact)",
    "Compute": "Low (Underutilized)",
    "Storage": "Low (No Downtime)",
    "Migration": "Low (No Active Tasks)",
    "AI/ML": "Medium (Model Serving Impact)",
}


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_len + 3


def generate_report(
    account_id: str,
    findings: list[Finding],
    mtd_by_service: dict[str, float],
    mtd_total: float,
    l3m_by_service_month: dict[str, dict[str, float]],
    l3m_months: list[str],
    projected_monthly: float,
    regions: list[str],
    scan_errors: list[str],
) -> Path:
    """Generate the 4-sheet Excel report. Returns the file path."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    total_savings = sum(f.savings for f in findings)
    target_spend = projected_monthly - total_savings
    reduction_pct = (total_savings / projected_monthly * 100) if projected_monthly > 0 else 0

    # Category breakdown
    categories: dict[str, dict] = {}
    for f in findings:
        if f.category not in categories:
            categories[f.category] = {"waste": 0.0, "savings": 0.0, "priority": f.priority}
        categories[f.category]["waste"] += f.savings
        categories[f.category]["savings"] += f.savings

    wb = Workbook()

    # --- Sheet 1: Executive Summary ---
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = "1F4E79"
    ws1["A1"] = "AWS FinOps — Cost Optimization Report"
    ws1["A1"].font = _TITLE_FONT
    ws1.merge_cells("A1:F1")

    for i, h in enumerate(["AWS Account ID", "Report Date", "Monthly Spend (Projected)", "Target Monthly Spend", "Total Potential Savings", "Reduction %"], 1):
        ws1.cell(row=3, column=i, value=h)
    _style_header_row(ws1, 3, 6)
    ws1["A4"] = account_id
    ws1["B4"] = today
    ws1["C4"] = projected_monthly
    ws1["C4"].number_format = _MONEY_FMT
    ws1["D4"] = target_spend
    ws1["D4"].number_format = _MONEY_FMT
    ws1["E4"] = total_savings
    ws1["E4"].number_format = _MONEY_FMT
    ws1["F4"] = f"{reduction_pct:.1f}%"

    ws1["A6"] = "Category Breakdown"
    ws1["A6"].font = Font(bold=True, size=12)
    for i, h in enumerate(["Resource Category", "Monthly Waste ($)", "Potential Savings ($)", "Priority", "Risk Impact"], 1):
        ws1.cell(row=7, column=i, value=h)
    _style_header_row(ws1, 7, 5)
    row = 8
    for cat, data in categories.items():
        ws1.cell(row=row, column=1, value=cat)
        ws1.cell(row=row, column=2, value=data["waste"]).number_format = _MONEY_FMT
        ws1.cell(row=row, column=3, value=data["savings"]).number_format = _MONEY_FMT
        cell = ws1.cell(row=row, column=4, value=data["priority"])
        cell.fill = _PRIORITY_FILLS.get(data["priority"], PatternFill())
        ws1.cell(row=row, column=5, value=_RISK_MAP.get(cat, "Low"))
        for col in range(1, 6):
            ws1.cell(row=row, column=col).border = _THIN_BORDER
        row += 1

    row += 1
    ws1.cell(row=row, column=1, value="Scan Coverage").font = Font(bold=True, size=12)
    row += 1
    ws1.cell(row=row, column=1, value=f"Regions scanned: {', '.join(regions)}")
    row += 1
    ws1.cell(row=row, column=1, value="Resource types checked: 20")
    row += 1
    ws1.cell(row=row, column=1, value=f"API errors encountered: {len(scan_errors)}")
    if scan_errors:
        row += 1
        ws1.cell(row=row, column=1, value="Failed checks:").font = Font(bold=True)
        for err_msg in scan_errors[:10]:
            row += 1
            ws1.cell(row=row, column=1, value=f"  ⚠ {err_msg}")
    _auto_width(ws1)

    # --- Sheet 2: Month-to-Date Costs ---
    mtd_start = datetime.now(timezone.utc).strftime("%Y-%m-01")
    days_in_month = max((datetime.now(timezone.utc) - datetime.strptime(mtd_start, "%Y-%m-%d").replace(tzinfo=timezone.utc)).days, 1)

    ws_mtd = wb.create_sheet("Cost - Month-to-Date")
    ws_mtd.sheet_properties.tabColor = "2E75B6"
    ws_mtd["A1"] = f"AWS Cost Breakdown — Month-to-Date ({mtd_start} to {today})"
    ws_mtd["A1"].font = _TITLE_FONT
    ws_mtd.merge_cells("A1:D1")
    ws_mtd["A2"] = f"Account: {account_id}"
    ws_mtd["A2"].font = Font(italic=True, size=10)
    for i, h in enumerate(["Service", "MTD Cost ($)", "% of Total", "Daily Average ($)"], 1):
        ws_mtd.cell(row=4, column=i, value=h)
    _style_header_row(ws_mtd, 4, 4)
    row = 5
    for svc, cost in sorted(mtd_by_service.items(), key=lambda x: x[1], reverse=True):
        if cost < 0.01:
            continue
        ws_mtd.cell(row=row, column=1, value=svc)
        ws_mtd.cell(row=row, column=2, value=round(cost, 2)).number_format = _MONEY_FMT
        pct = (cost / mtd_total * 100) if mtd_total > 0 else 0
        ws_mtd.cell(row=row, column=3, value=f"{pct:.1f}%")
        ws_mtd.cell(row=row, column=4, value=round(cost / days_in_month, 2)).number_format = _MONEY_FMT
        for col in range(1, 5):
            ws_mtd.cell(row=row, column=col).border = _THIN_BORDER
        row += 1
    # Total row
    ws_mtd.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws_mtd.cell(row=row, column=2, value=round(mtd_total, 2)).number_format = _MONEY_FMT
    ws_mtd.cell(row=row, column=2).font = Font(bold=True)
    ws_mtd.cell(row=row, column=3, value="100%").font = Font(bold=True)
    ws_mtd.cell(row=row, column=4, value=round(mtd_total / days_in_month, 2)).number_format = _MONEY_FMT
    ws_mtd.cell(row=row, column=4).font = Font(bold=True)
    for col in range(1, 5):
        ws_mtd.cell(row=row, column=col).border = _THIN_BORDER
    _auto_width(ws_mtd)

    # --- Sheet 3: Last 3 Months ---
    ws_l3m = wb.create_sheet("Cost - Last 3 Months")
    ws_l3m.sheet_properties.tabColor = "70AD47"
    ws_l3m["A1"] = "AWS Cost Breakdown — Last 3 Months"
    ws_l3m["A1"].font = _TITLE_FONT
    headers_l3m = ["Service"] + l3m_months + ["Total ($)", "Avg Monthly ($)"]
    col_count = len(headers_l3m)
    if col_count > 1:
        ws_l3m.merge_cells(f"A1:{get_column_letter(col_count)}1")
    ws_l3m["A2"] = f"Account: {account_id}"
    ws_l3m["A2"].font = Font(italic=True, size=10)
    for i, h in enumerate(headers_l3m, 1):
        ws_l3m.cell(row=4, column=i, value=h)
    _style_header_row(ws_l3m, 4, col_count)

    svc_totals = {svc: sum(months.values()) for svc, months in l3m_by_service_month.items()}
    row = 5
    grand_totals = {m: 0.0 for m in l3m_months}
    grand_total = 0.0
    for svc, total in sorted(svc_totals.items(), key=lambda x: x[1], reverse=True):
        if total < 0.01:
            continue
        ws_l3m.cell(row=row, column=1, value=svc)
        for mi, month in enumerate(l3m_months, 2):
            amt = l3m_by_service_month[svc].get(month, 0)
            ws_l3m.cell(row=row, column=mi, value=round(amt, 2)).number_format = _MONEY_FMT
            grand_totals[month] += amt
        ws_l3m.cell(row=row, column=len(l3m_months) + 2, value=round(total, 2)).number_format = _MONEY_FMT
        ws_l3m.cell(row=row, column=len(l3m_months) + 3, value=round(total / max(len(l3m_months), 1), 2)).number_format = _MONEY_FMT
        grand_total += total
        for col in range(1, col_count + 1):
            ws_l3m.cell(row=row, column=col).border = _THIN_BORDER
        row += 1
    # Total row
    ws_l3m.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for mi, month in enumerate(l3m_months, 2):
        ws_l3m.cell(row=row, column=mi, value=round(grand_totals[month], 2)).number_format = _MONEY_FMT
        ws_l3m.cell(row=row, column=mi).font = Font(bold=True)
    ws_l3m.cell(row=row, column=len(l3m_months) + 2, value=round(grand_total, 2)).number_format = _MONEY_FMT
    ws_l3m.cell(row=row, column=len(l3m_months) + 2).font = Font(bold=True)
    ws_l3m.cell(row=row, column=len(l3m_months) + 3, value=round(grand_total / max(len(l3m_months), 1), 2)).number_format = _MONEY_FMT
    ws_l3m.cell(row=row, column=len(l3m_months) + 3).font = Font(bold=True)
    for col in range(1, col_count + 1):
        ws_l3m.cell(row=row, column=col).border = _THIN_BORDER
    _auto_width(ws_l3m)

    # --- Sheet 4: Remediation Ledger ---
    ws2 = wb.create_sheet("Remediation Ledger")
    ws2.sheet_properties.tabColor = "C00000"
    ws2["A1"] = "Technical Remediation Ledger"
    ws2["A1"].font = _TITLE_FONT
    ws2.merge_cells("A1:H1")
    headers2 = ["Resource ID", "Resource Type", "Region", "Current Configuration",
                "Suggested Remediation", "Est. Monthly Savings ($)", "Reasoning", "Remediation Command"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=i, value=h)
    _style_header_row(ws2, 3, 8)
    for idx, f in enumerate(findings, 4):
        ws2.cell(row=idx, column=1, value=f.resource_id)
        ws2.cell(row=idx, column=2, value=f.resource_type)
        ws2.cell(row=idx, column=3, value=f.region)
        ws2.cell(row=idx, column=4, value=f.config)
        ws2.cell(row=idx, column=5, value=f.remediation)
        ws2.cell(row=idx, column=6, value=f.savings).number_format = _MONEY_FMT
        ws2.cell(row=idx, column=7, value=f.reasoning)
        ws2.cell(row=idx, column=8, value=f.command)
        for col in range(1, 9):
            ws2.cell(row=idx, column=col).border = _THIN_BORDER
            ws2.cell(row=idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    _auto_width(ws2)

    # Save
    filename = f"finops-report-{account_id}-{today}.xlsx"
    filepath = LOG_DIR / filename
    wb.save(filepath)
    return filepath


def generate_org_report(
    mgmt_account_id: str,
    account_results: list[dict],
    scan_errors: list[str],
    regions: list[str],
) -> Path:
    """Generate organization-wide report with per-account tabs."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    tab_colors = ["1F4E79", "2E75B6", "70AD47", "C00000", "7030A0", "ED7D31", "44546A", "4472C4"]

    wb = Workbook()
    wb.remove(wb.active)

    # Org Summary sheet
    ws_summary = wb.create_sheet("Org Summary", 0)
    ws_summary.sheet_properties.tabColor = "000000"
    ws_summary["A1"] = "AWS Organization — Cost Optimization Summary"
    ws_summary["A1"].font = _TITLE_FONT
    ws_summary.merge_cells("A1:E1")
    ws_summary["A2"] = f"Management Account: {mgmt_account_id} | Date: {today} | Regions: {', '.join(regions)}"
    ws_summary["A2"].font = Font(italic=True, size=10)

    headers = ["Account Name", "Account ID", "Findings", "Est. Monthly Savings ($)", "Status"]
    for i, h in enumerate(headers, 1):
        ws_summary.cell(row=4, column=i, value=h)
    _style_header_row(ws_summary, 4, 5)

    row = 5
    total_org_savings = 0.0
    total_findings = 0
    for acct in account_results:
        ws_summary.cell(row=row, column=1, value=acct["name"])
        ws_summary.cell(row=row, column=2, value=acct["id"])
        ws_summary.cell(row=row, column=3, value=acct["finding_count"])
        ws_summary.cell(row=row, column=4, value=acct["savings"]).number_format = _MONEY_FMT
        ws_summary.cell(row=row, column=5, value="Scanned ✓")
        total_org_savings += acct["savings"]
        total_findings += acct["finding_count"]
        for col in range(1, 6):
            ws_summary.cell(row=row, column=col).border = _THIN_BORDER
        row += 1

    # Failed accounts
    for err_msg in scan_errors:
        if "Cannot assume role" in err_msg:
            ws_summary.cell(row=row, column=1, value=err_msg.split(":")[0])
            ws_summary.cell(row=row, column=5, value="⚠ Access Denied")
            ws_summary.cell(row=row, column=5).font = Font(color="FF0000")
            for col in range(1, 6):
                ws_summary.cell(row=row, column=col).border = _THIN_BORDER
            row += 1

    # Total
    ws_summary.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=row, column=3, value=total_findings).font = Font(bold=True)
    ws_summary.cell(row=row, column=4, value=total_org_savings).number_format = _MONEY_FMT
    ws_summary.cell(row=row, column=4).font = Font(bold=True)
    for col in range(1, 6):
        ws_summary.cell(row=row, column=col).border = _THIN_BORDER
    _auto_width(ws_summary)

    # Per-account sheets
    for idx, acct in enumerate(account_results):
        sheet_name = f"{acct['name'][:20]} ({acct['id'][-4:]})"
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab_colors[idx % len(tab_colors)]
        ws["A1"] = f"Account: {acct['name']} ({acct['id']})"
        ws["A1"].font = _TITLE_FONT
        ws.merge_cells("A1:F1")

        headers = ["Resource ID", "Type", "Region", "Configuration", "Suggested Remediation", "Est. Savings ($)"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        _style_header_row(ws, 3, 6)

        findings = acct.get("findings", [])
        for fi, f in enumerate(findings, 4):
            ws.cell(row=fi, column=1, value=f.resource_id)
            ws.cell(row=fi, column=2, value=f.resource_type)
            ws.cell(row=fi, column=3, value=f.region)
            ws.cell(row=fi, column=4, value=f.config)
            ws.cell(row=fi, column=5, value=f.remediation)
            ws.cell(row=fi, column=6, value=f.savings).number_format = _MONEY_FMT
            for col in range(1, 7):
                ws.cell(row=fi, column=col).border = _THIN_BORDER

        if not findings:
            ws.cell(row=4, column=1, value="No optimization opportunities found.")
        _auto_width(ws)

    filename = f"finops-org-report-{mgmt_account_id}-{today}.xlsx"
    filepath = LOG_DIR / filename
    wb.save(filepath)
    return filepath


def save_history(account_id: str, findings: list[Finding], mtd_total: float, projected_monthly: float):
    """Append run results to history JSON for trend tracking."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    total_savings = sum(f.savings for f in findings)
    history_file = LOG_DIR / f"history-{account_id}.json"
    history = []
    if history_file.exists():
        try:
            history = json.loads(history_file.read_text())
        except (json.JSONDecodeError, OSError):
            history = []
    history.append({
        "date": today,
        "mtd_spend": round(mtd_total, 2),
        "projected_monthly": round(projected_monthly, 2),
        "total_savings_identified": round(total_savings, 2),
        "finding_count": len(findings),
        "findings_by_priority": {
            "HIGH": sum(1 for f in findings if f.priority == "HIGH"),
            "MEDIUM": sum(1 for f in findings if f.priority == "MEDIUM"),
            "LOW": sum(1 for f in findings if f.priority == "LOW"),
        },
    })
    history_file.write_text(json.dumps(history, indent=2))
