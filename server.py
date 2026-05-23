import json
import logging
import re
import shlex
import subprocess
from datetime import datetime, timedelta
from pathlib import Path

from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

mcp = FastMCP("finops-mcp-agent")

# --- Logging ---
LOG_DIR = Path.home() / "finops-reports"
LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    filename=str(LOG_DIR / "finops-agent.log"),
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("finops")

# --- Allowed remediation command patterns (whitelist) ---
ALLOWED_COMMANDS = [
    r"^aws ec2 (terminate-instances|stop-instances|delete-volume|release-address|delete-nat-gateway|delete-vpn-connection|delete-client-vpn-endpoint|delete-snapshot) ",
    r"^aws rds delete-db-instance ",
    r"^aws elbv2 delete-load-balancer ",
    r"^aws elb delete-load-balancer ",
    r"^aws elasticache delete-cache-cluster ",
    r"^aws opensearch delete-domain ",
    r"^aws redshift delete-cluster ",
    r"^aws dms delete-replication-instance ",
    r"^aws logs put-retention-policy ",
    r"^aws ecs delete-service ",
    r"^aws eks (delete-cluster|delete-nodegroup|update-nodegroup-config) ",
    r"^aws sagemaker (delete-endpoint|stop-notebook-instance|delete-notebook-instance|stop-training-job) ",
    r"^aws bedrock delete-provisioned-model-throughput ",
]


def validate_command(command: str) -> bool:
    """Check if a remediation command matches the whitelist."""
    return any(re.match(pattern, command) for pattern in ALLOWED_COMMANDS)


def aws_cmd(cmd: str, profile: str, region: str) -> tuple:
    """Execute an AWS CLI command safely. Returns (data, error_msg).
    If profile is empty/None, uses environment credentials (AWS_ACCESS_KEY_ID etc)."""
    args = ["aws"] + shlex.split(cmd) + ["--region", region, "--output", "json"]
    if profile:
        args += ["--profile", profile]
    try:
        result = subprocess.run(args, capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            data = json.loads(result.stdout) if result.stdout.strip() else None
            return data, None
        else:
            err = result.stderr.strip()[:200]
            logger.warning(f"AWS CLI failed: {' '.join(args[:5])}... -> {err}")
            return None, err
    except subprocess.TimeoutExpired:
        logger.error(f"Timeout: {' '.join(args[:5])}...")
        return None, "Command timed out (60s)"
    except json.JSONDecodeError as e:
        logger.error(f"JSON parse error: {e}")
        return None, "Invalid JSON response"
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return None, str(e)


def aws_cmd_with_role(cmd: str, role_arn: str, region: str, session_name: str = "finops-scan") -> tuple:
    """Execute AWS CLI command by assuming a role. For cross-account scanning."""
    # First assume the role
    assume_args = [
        "aws", "sts", "assume-role",
        "--role-arn", role_arn,
        "--role-session-name", session_name,
        "--output", "json"
    ]
    try:
        assume_result = subprocess.run(assume_args, capture_output=True, text=True, timeout=30)
        if assume_result.returncode != 0:
            return None, f"Cannot assume role {role_arn}: {assume_result.stderr.strip()[:100]}"
        creds = json.loads(assume_result.stdout).get("Credentials", {})
    except Exception as e:
        return None, f"Role assumption failed: {e}"

    # Execute the actual command with assumed credentials
    env_override = {
        "AWS_ACCESS_KEY_ID": creds["AccessKeyId"],
        "AWS_SECRET_ACCESS_KEY": creds["SecretAccessKey"],
        "AWS_SESSION_TOKEN": creds["SessionToken"],
    }
    import os
    env = {**os.environ, **env_override}
    args = ["aws"] + shlex.split(cmd) + ["--region", region, "--output", "json"]
    try:
        result = subprocess.run(args, capture_output=True, text=True, timeout=60, env=env)
        if result.returncode == 0:
            data = json.loads(result.stdout) if result.stdout.strip() else None
            return data, None
        else:
            return None, result.stderr.strip()[:200]
    except subprocess.TimeoutExpired:
        return None, "Command timed out (60s)"
    except Exception as e:
        return None, str(e)


def get_metric(profile, region, namespace, metric, dim_name, dim_value, start, end):
    data, _ = aws_cmd(
        f"cloudwatch get-metric-statistics --namespace {namespace} --metric-name {metric} "
        f"--dimensions Name={dim_name},Value={dim_value} "
        f"--start-time {start}T00:00:00Z --end-time {end}T00:00:00Z --period 86400 --statistics Average",
        profile, region
    )
    return data


# --- Pricing Lookup ---
# On-demand hourly rates (USD) for common instance types.
# Used when Pricing API is unavailable or too slow.
# Source: AWS pricing pages, updated periodically.
REGION_NAMES = {
    "us-east-1": "US East (N. Virginia)", "us-east-2": "US East (Ohio)",
    "us-west-1": "US West (N. California)", "us-west-2": "US West (Oregon)",
    "ap-south-1": "Asia Pacific (Mumbai)", "ap-southeast-1": "Asia Pacific (Singapore)",
    "ap-southeast-2": "Asia Pacific (Sydney)", "ap-northeast-1": "Asia Pacific (Tokyo)",
    "eu-west-1": "Europe (Ireland)", "eu-central-1": "Europe (Frankfurt)",
}

# EBS pricing per GB/month by volume type and region
EBS_PRICING = {
    "default": {"gp2": 0.10, "gp3": 0.08, "io1": 0.125, "io2": 0.125, "st1": 0.045, "sc1": 0.025, "standard": 0.05},
    "ap-south-1": {"gp2": 0.114, "gp3": 0.0912, "io1": 0.131, "io2": 0.131, "st1": 0.051, "sc1": 0.029, "standard": 0.057},
}


def get_ec2_hourly_price(instance_type: str, profile: str, region: str) -> float:
    """Get EC2 on-demand hourly price from Pricing API. Returns 0 on failure."""
    location = REGION_NAMES.get(region)
    if not location:
        return 0
    data, err = aws_cmd(
        f'pricing get-products --service-code AmazonEC2 '
        f'--filters "Type=TERM_MATCH,Field=instanceType,Value={instance_type}" '
        f'"Type=TERM_MATCH,Field=location,Value={location}" '
        f'"Type=TERM_MATCH,Field=operatingSystem,Value=Linux" '
        f'"Type=TERM_MATCH,Field=tenancy,Value=Shared" '
        f'"Type=TERM_MATCH,Field=preInstalledSw,Value=NA" '
        f'"Type=TERM_MATCH,Field=capacitystatus,Value=Used"',
        profile, "us-east-1"
    )
    if not data or not data.get("PriceList"):
        return 0
    try:
        price_item = json.loads(data["PriceList"][0])
        on_demand = price_item.get("terms", {}).get("OnDemand", {})
        for term in on_demand.values():
            for dim in term.get("priceDimensions", {}).values():
                price = float(dim.get("pricePerUnit", {}).get("USD", "0"))
                if price > 0:
                    return price
    except (json.JSONDecodeError, KeyError, IndexError, ValueError):
        pass
    return 0


def get_rds_hourly_price(instance_class: str, engine: str, profile: str, region: str) -> float:
    """Get RDS on-demand hourly price from Pricing API."""
    location = REGION_NAMES.get(region)
    if not location:
        return 0
    # Map engine to pricing API database engine value
    engine_map = {
        "aurora-postgresql": "Aurora PostgreSQL",
        "aurora-mysql": "Aurora MySQL",
        "postgres": "PostgreSQL",
        "mysql": "MySQL",
        "mariadb": "MariaDB",
        "oracle-ee": "Oracle",
        "sqlserver-ee": "SQL Server",
    }
    db_engine = engine_map.get(engine, engine)
    data, err = aws_cmd(
        f'pricing get-products --service-code AmazonRDS '
        f'--filters "Type=TERM_MATCH,Field=instanceType,Value={instance_class}" '
        f'"Type=TERM_MATCH,Field=location,Value={location}" '
        f'"Type=TERM_MATCH,Field=databaseEngine,Value={db_engine}" '
        f'"Type=TERM_MATCH,Field=deploymentOption,Value=Single-AZ"',
        profile, "us-east-1"
    )
    if not data or not data.get("PriceList"):
        return 0
    try:
        price_item = json.loads(data["PriceList"][0])
        on_demand = price_item.get("terms", {}).get("OnDemand", {})
        for term in on_demand.values():
            for dim in term.get("priceDimensions", {}).values():
                price = float(dim.get("pricePerUnit", {}).get("USD", "0"))
                if price > 0:
                    return price
    except (json.JSONDecodeError, KeyError, IndexError, ValueError):
        pass
    return 0


def get_ebs_monthly_cost(size_gb: int, vol_type: str, region: str) -> float:
    """Calculate EBS monthly cost using regional pricing."""
    pricing = EBS_PRICING.get(region, EBS_PRICING["default"])
    rate = pricing.get(vol_type, 0.08)
    return round(size_gb * rate, 2)


def save_history(account_id, today, findings, mtd_total, projected_monthly, total_savings):
    """Append run results to history JSON for trend tracking."""
    history_file = LOG_DIR / f"history-{account_id}.json"
    history = []
    if history_file.exists():
        try:
            history = json.loads(history_file.read_text())
        except (json.JSONDecodeError, OSError):
            history = []
    history.append({
        "date": today,
        "mtd_spend": round(mtd_total, 2),
        "projected_monthly": round(projected_monthly, 2),
        "total_savings_identified": round(total_savings, 2),
        "finding_count": len(findings),
        "findings_by_priority": {
            "HIGH": sum(1 for f in findings if f["priority"] == "HIGH"),
            "MEDIUM": sum(1 for f in findings if f["priority"] == "MEDIUM"),
            "LOW": sum(1 for f in findings if f["priority"] == "LOW"),
        },
    })
    history_file.write_text(json.dumps(history, indent=2))


@mcp.tool()
def analyze_aws_costs(profile: str, regions: list[str], lookback_days: int = 30) -> str:
    """
    Analyze AWS account for cost optimization opportunities.
    Scans 20+ resource types (EC2, RDS, EBS, ELB, ElastiCache, OpenSearch,
    Redshift, CloudWatch Logs, ECS, EKS, SageMaker, Bedrock, Elastic IPs,
    VPN, NAT, DMS) and produces a styled Excel report with:
    - Executive summary with projected spend and savings
    - Month-to-date cost breakdown by service (from Cost Explorer)
    - Last 3 months cost trends by service (from Cost Explorer)
    - Detailed remediation ledger with runnable commands

    Args:
        profile: AWS CLI profile name (e.g. 'rudraksha')
        regions: List of AWS regions to scan (e.g. ['us-east-1', 'ap-south-1'])
        lookback_days: Number of days to look back for utilization metrics (default: 30)
    """
    # Input validation
    if not profile or not profile.isalnum() and not all(c.isalnum() or c in "-_" for c in profile):
        return "❌ Invalid profile name. Use alphanumeric characters, hyphens, or underscores."
    region_pattern = re.compile(r"^[a-z]{2}-[a-z]+-\d+$")
    for r in regions:
        if not region_pattern.match(r):
            return f"❌ Invalid region format: '{r}'. Expected format: us-east-1, ap-south-1, etc."

    logger.info(f"Starting analysis: profile={profile}, regions={regions}, lookback={lookback_days}")
    scan_errors = []  # Track failed API calls

    end_date = datetime.utcnow().strftime("%Y-%m-%d")
    start_date = (datetime.utcnow() - timedelta(days=lookback_days)).strftime("%Y-%m-%d")
    today = end_date

    # Get account ID
    identity, err = aws_cmd("sts get-caller-identity", profile, regions[0])
    if err:
        return f"❌ Cannot authenticate with profile '{profile}': {err}"
    account_id = identity.get("Account", "unknown")

    # --- Cost Explorer: Month-to-Date ---
    mtd_start = datetime.utcnow().strftime("%Y-%m-01")
    mtd_data, err = aws_cmd(
        f"ce get-cost-and-usage --granularity DAILY --group-by Type=DIMENSION,Key=SERVICE "
        f"--metrics UnblendedCost --time-period Start={mtd_start},End={end_date}",
        profile, "us-east-1"
    )
    if err:
        scan_errors.append(f"Cost Explorer MTD: {err}")

    # --- Cost Explorer: Last 3 months ---
    three_months_ago = (datetime.utcnow().replace(day=1) - timedelta(days=90)).strftime("%Y-%m-01")
    last3m_data, err = aws_cmd(
        f"ce get-cost-and-usage --granularity MONTHLY --group-by Type=DIMENSION,Key=SERVICE "
        f"--metrics UnblendedCost --time-period Start={three_months_ago},End={mtd_start}",
        profile, "us-east-1"
    )
    if err:
        scan_errors.append(f"Cost Explorer 3-month: {err}")

    # Parse MTD costs
    mtd_by_service = {}
    mtd_total = 0.0
    if mtd_data and "ResultsByTime" in mtd_data:
        for period in mtd_data["ResultsByTime"]:
            for group in period.get("Groups", []):
                svc = group["Keys"][0]
                amt = float(group.get("Metrics", {}).get("UnblendedCost", {}).get("Amount", "0"))
                mtd_by_service[svc] = mtd_by_service.get(svc, 0) + amt
                mtd_total += amt

    # Parse last 3 months
    l3m_by_service_month = {}
    l3m_months = []
    if last3m_data and "ResultsByTime" in last3m_data:
        for period in last3m_data["ResultsByTime"]:
            month_label = period["TimePeriod"]["Start"][:7]
            if month_label not in l3m_months:
                l3m_months.append(month_label)
            for group in period.get("Groups", []):
                svc = group["Keys"][0]
                amt = float(group.get("Metrics", {}).get("UnblendedCost", {}).get("Amount", "0"))
                if svc not in l3m_by_service_month:
                    l3m_by_service_month[svc] = {}
                l3m_by_service_month[svc][month_label] = l3m_by_service_month[svc].get(month_label, 0) + amt

    days_in_month_so_far = max((datetime.utcnow() - datetime.strptime(mtd_start, "%Y-%m-%d")).days, 1)
    projected_monthly = (mtd_total / days_in_month_so_far) * 30

    findings = []

    for region in regions:
        # --- EC2 Instances ---
        instances, err = aws_cmd(
            'ec2 describe-instances --query "Reservations[].Instances[].{InstanceId:InstanceId,InstanceType:InstanceType,State:State.Name,Tags:Tags}"',
            profile, region
        )
        if err:
            scan_errors.append(f"EC2/{region}: {err}")
            instances = []
        instances = instances or []
        for inst in instances:
            if inst["State"] == "stopped":
                findings.append({
                    "id": inst["InstanceId"], "type": "EC2 Instance (Stopped)", "region": region,
                    "config": f"{inst['InstanceType']} (stopped)",
                    "remediation": "Terminate stopped instance to eliminate EBS charges",
                    "savings": 0.80,
                    "reasoning": "Stopped instance still incurs EBS volume charges. Terminate if not needed.",
                    "command": f"aws ec2 terminate-instances --instance-ids {inst['InstanceId']} --profile {profile} --region {region}",
                    "category": "Compute", "priority": "MEDIUM"
                })
            elif inst["State"] == "running":
                cpu = get_metric(profile, region, "AWS/EC2", "CPUUtilization", "InstanceId", inst["InstanceId"], start_date, end_date)
                if cpu and cpu.get("Datapoints"):
                    avg = sum(d["Average"] for d in cpu["Datapoints"]) / len(cpu["Datapoints"])
                    if avg < 5:
                        hourly = get_ec2_hourly_price(inst["InstanceType"], profile, region)
                        monthly_cost = round(hourly * 730, 2) if hourly else 8.00
                        findings.append({
                            "id": inst["InstanceId"], "type": "EC2 Instance (Underutilized)", "region": region,
                            "config": f"{inst['InstanceType']} (avg CPU: {avg:.2f}%) — ${monthly_cost}/mo on-demand",
                            "remediation": "Downsize or terminate - CPU under 5%",
                            "savings": monthly_cost,
                            "reasoning": f"Average CPU utilization is {avg:.2f}% over {len(cpu['Datapoints'])} days. Significantly over-provisioned.",
                            "command": f"aws ec2 stop-instances --instance-ids {inst['InstanceId']} --profile {profile} --region {region}",
                            "category": "Compute", "priority": "MEDIUM"
                        })

        # --- EBS Unattached Volumes ---
        volumes, err = aws_cmd(
            'ec2 describe-volumes --query "Volumes[].{VolumeId:VolumeId,Size:Size,VolumeType:VolumeType,State:State,Attachments:Attachments}"',
            profile, region
        )
        if err:
            scan_errors.append(f"EBS/{region}: {err}")
            volumes = []
        volumes = volumes or []
        for vol in volumes:
            if vol["State"] == "available" and not vol.get("Attachments"):
                cost = get_ebs_monthly_cost(vol["Size"], vol["VolumeType"], region)
                findings.append({
                    "id": vol["VolumeId"], "type": "EBS Volume (Unattached)", "region": region,
                    "config": f"{vol['Size']}GB {vol['VolumeType']} - unattached",
                    "remediation": "Delete unattached volume",
                    "savings": cost,
                    "reasoning": "Volume in 'available' state with no attachments. Incurring storage charges with no use.",
                    "command": f"aws ec2 delete-volume --volume-id {vol['VolumeId']} --profile {profile} --region {region}",
                    "category": "Storage", "priority": "MEDIUM"
                })

        # --- RDS Instances (with replica awareness) ---
        rds_instances, err = aws_cmd(
            'rds describe-db-instances --query "DBInstances[].{DBInstanceIdentifier:DBInstanceIdentifier,DBInstanceClass:DBInstanceClass,Engine:Engine,AllocatedStorage:AllocatedStorage,DBClusterIdentifier:DBClusterIdentifier,ReadReplicaSourceDBInstanceIdentifier:ReadReplicaSourceDBInstanceIdentifier}"',
            profile, region
        )
        if err:
            scan_errors.append(f"RDS/{region}: {err}")
            rds_instances = []
        rds_instances = rds_instances or []
        for db in rds_instances:
            # Skip read replicas - they legitimately have zero direct connections
            is_replica = bool(db.get("ReadReplicaSourceDBInstanceIdentifier"))
            is_cluster_member = bool(db.get("DBClusterIdentifier"))

            conns = get_metric(profile, region, "AWS/RDS", "DatabaseConnections", "DBInstanceIdentifier", db["DBInstanceIdentifier"], start_date, end_date)
            if conns and conns.get("Datapoints"):
                avg_conns = sum(d["Average"] for d in conns["Datapoints"]) / len(conns["Datapoints"])
                if avg_conns < 1 and not is_replica:
                    hourly = get_rds_hourly_price(db["DBInstanceClass"], db["Engine"], profile, region)
                    est_cost = round(hourly * 730, 2) if hourly else (52 if "large" in db["DBInstanceClass"] else 26)
                    context = ""
                    if is_cluster_member:
                        context = f" (Aurora cluster member: {db['DBClusterIdentifier']})"
                    findings.append({
                        "id": db["DBInstanceIdentifier"], "type": f"RDS {db['Engine']}", "region": region,
                        "config": f"{db['DBInstanceClass']} ({db['Engine']} / {db['AllocatedStorage']}GB){context}",
                        "remediation": "Terminate - zero connections" if avg_conns == 0 else "Evaluate - near-zero connections",
                        "savings": est_cost,
                        "reasoning": f"Average connections = {avg_conns:.2f} over analysis period. Instance appears idle.{' Note: This is an Aurora cluster member - evaluate cluster-wide.' if is_cluster_member else ''}",
                        "command": f"aws rds delete-db-instance --db-instance-identifier {db['DBInstanceIdentifier']} --skip-final-snapshot --profile {profile} --region {region}",
                        "category": "Database", "priority": "HIGH"
                    })

        # --- NAT Gateways ---
        nats, err = aws_cmd('ec2 describe-nat-gateways --query "NatGateways[].{NatGatewayId:NatGatewayId,State:State,VpcId:VpcId}"', profile, region)
        if err:
            scan_errors.append(f"NAT/{region}: {err}")
        for nat in (nats or []):
            if nat["State"] == "available":
                findings.append({
                    "id": nat["NatGatewayId"], "type": "NAT Gateway", "region": region,
                    "config": f"Active NAT Gateway in {nat['VpcId']}",
                    "remediation": "Evaluate necessity - costs ~$32/mo minimum",
                    "savings": 5.18,
                    "reasoning": "NAT Gateway costs $0.045/hr + data processing. Evaluate if private subnet egress is required.",
                    "command": f"aws ec2 delete-nat-gateway --nat-gateway-id {nat['NatGatewayId']} --profile {profile} --region {region}",
                    "category": "Networking", "priority": "MEDIUM"
                })

        # --- DMS ---
        dms_instances, err = aws_cmd(
            'dms describe-replication-instances --query "ReplicationInstances[].{ReplicationInstanceIdentifier:ReplicationInstanceIdentifier,ReplicationInstanceClass:ReplicationInstanceClass,ReplicationInstanceStatus:ReplicationInstanceStatus,ReplicationInstanceArn:ReplicationInstanceArn}"',
            profile, region
        )
        if err:
            scan_errors.append(f"DMS/{region}: {err}")
        dms_tasks, _ = aws_cmd('dms describe-replication-tasks --query "ReplicationTasks[].{Status:Status}"', profile, region)
        for dms in (dms_instances or []):
            if dms["ReplicationInstanceStatus"] == "available" and not (dms_tasks or []):
                findings.append({
                    "id": dms["ReplicationInstanceIdentifier"], "type": "DMS Replication Instance", "region": region,
                    "config": f"{dms['ReplicationInstanceClass']} (no active tasks)",
                    "remediation": "Terminate idle DMS instance",
                    "savings": 6.80,
                    "reasoning": "Replication instance running with zero active tasks. Billed continuously with no workload.",
                    "command": f"aws dms delete-replication-instance --replication-instance-arn {dms['ReplicationInstanceArn']} --profile {profile} --region {region}",
                    "category": "Migration", "priority": "LOW"
                })

        # --- Snapshots ---
        snapshots, err = aws_cmd('ec2 describe-snapshots --owner-ids self --query "Snapshots[].{SnapshotId:SnapshotId,VolumeSize:VolumeSize,Description:Description}"', profile, region)
        if err:
            scan_errors.append(f"Snapshots/{region}: {err}")
        migration_snaps = [s for s in (snapshots or []) if "Application Migration Service" in (s.get("Description") or "")]
        if migration_snaps:
            total_size = sum(s["VolumeSize"] for s in migration_snaps)
            cost = round(total_size * 0.05 / 10, 2)
            snap_ids = " ".join(s["SnapshotId"] for s in migration_snaps)
            findings.append({
                "id": f"{len(migration_snaps)} snapshots", "type": f"EBS Snapshots ({len(migration_snaps)} orphaned)", "region": region,
                "config": f"{len(migration_snaps)} migration snapshots (~{total_size}GB)",
                "remediation": "Delete orphaned migration snapshots",
                "savings": cost,
                "reasoning": "Migration snapshots from completed migration. No longer needed.",
                "command": f"aws ec2 delete-snapshot --snapshot-id {migration_snaps[0]['SnapshotId']} --profile {profile} --region {region}",
                "category": "Storage", "priority": "LOW"
            })

        # --- VPN ---
        vpns, err = aws_cmd('ec2 describe-vpn-connections --query "VpnConnections[].{VpnConnectionId:VpnConnectionId,State:State,Type:Type}"', profile, region)
        if err:
            scan_errors.append(f"VPN/{region}: {err}")
        for vpn in (vpns or []):
            if vpn["State"] == "available":
                findings.append({
                    "id": vpn["VpnConnectionId"], "type": "Site-to-Site VPN", "region": region,
                    "config": f"IPSec VPN ({vpn['Type']})",
                    "remediation": "Evaluate necessity",
                    "savings": 23.50,
                    "reasoning": "VPN connection costs ~$0.048/hr. Evaluate if on-premises connectivity is actively required.",
                    "command": f"aws ec2 delete-vpn-connection --vpn-connection-id {vpn['VpnConnectionId']} --profile {profile} --region {region}",
                    "category": "Networking", "priority": "HIGH"
                })

        # --- Client VPN ---
        cvpns, err = aws_cmd('ec2 describe-client-vpn-endpoints --query "ClientVpnEndpoints[].{ClientVpnEndpointId:ClientVpnEndpointId,Status:Status}"', profile, region)
        if err:
            scan_errors.append(f"ClientVPN/{region}: {err}")
        for cvpn in (cvpns or []):
            if cvpn.get("Status", {}).get("Code") == "pending-associate":
                findings.append({
                    "id": cvpn["ClientVpnEndpointId"], "type": "Client VPN Endpoint", "region": region,
                    "config": "Status: pending-associate (no subnet associations)",
                    "remediation": "Delete unused Client VPN endpoint",
                    "savings": 24.00,
                    "reasoning": "Client VPN in 'pending-associate' state. Being billed but serving no traffic.",
                    "command": f"aws ec2 delete-client-vpn-endpoint --client-vpn-endpoint-id {cvpn['ClientVpnEndpointId']} --profile {profile} --region {region}",
                    "category": "Networking", "priority": "HIGH"
                })

        # --- Unassociated Elastic IPs ---
        eips, err = aws_cmd('ec2 describe-addresses --query "Addresses[].{AllocationId:AllocationId,PublicIp:PublicIp,AssociationId:AssociationId}"', profile, region)
        if err:
            scan_errors.append(f"EIP/{region}: {err}")
        for eip in (eips or []):
            if not eip.get("AssociationId"):
                findings.append({
                    "id": eip["AllocationId"], "type": "Elastic IP (Unassociated)", "region": region,
                    "config": f"{eip['PublicIp']} - not associated to any resource",
                    "remediation": "Release unassociated Elastic IP",
                    "savings": 3.60,
                    "reasoning": "Unassociated EIPs cost $0.005/hr ($3.60/mo). Release if not needed.",
                    "command": f"aws ec2 release-address --allocation-id {eip['AllocationId']} --profile {profile} --region {region}",
                    "category": "Networking", "priority": "LOW"
                })

        # --- Idle Load Balancers (ALB/NLB) ---
        lbs, err = aws_cmd('elbv2 describe-load-balancers --query "LoadBalancers[].{LoadBalancerArn:LoadBalancerArn,LoadBalancerName:LoadBalancerName,Type:Type,State:State.Code}"', profile, region)
        if err:
            scan_errors.append(f"ELBv2/{region}: {err}")
        for lb in (lbs or []):
            if lb.get("State") != "active":
                continue
            tgs, _ = aws_cmd(f'elbv2 describe-target-groups --load-balancer-arn {lb["LoadBalancerArn"]} --query "TargetGroups[].TargetGroupArn"', profile, region)
            has_targets = False
            for tg_arn in (tgs or []):
                health, _ = aws_cmd(f'elbv2 describe-target-health --target-group-arn {tg_arn} --query "TargetHealthDescriptions[]"', profile, region)
                if health:
                    has_targets = True
                    break
            if not has_targets:
                findings.append({
                    "id": lb["LoadBalancerName"], "type": f"Load Balancer ({lb['Type'].upper()})", "region": region,
                    "config": f"{lb['Type']} LB with no healthy targets",
                    "remediation": "Delete idle load balancer",
                    "savings": 16.20,
                    "reasoning": "Load balancer has no healthy targets registered. Incurring hourly charges with no traffic served.",
                    "command": f"aws elbv2 delete-load-balancer --load-balancer-arn {lb['LoadBalancerArn']} --profile {profile} --region {region}",
                    "category": "Networking", "priority": "MEDIUM"
                })

        # --- Classic Load Balancers ---
        clbs, err = aws_cmd('elb describe-load-balancers --query "LoadBalancerDescriptions[].{LoadBalancerName:LoadBalancerName,Instances:Instances}"', profile, region)
        if err:
            scan_errors.append(f"CLB/{region}: {err}")
        for clb in (clbs or []):
            if not clb.get("Instances"):
                findings.append({
                    "id": clb["LoadBalancerName"], "type": "Classic Load Balancer (Idle)", "region": region,
                    "config": "Classic LB with no registered instances",
                    "remediation": "Delete idle Classic Load Balancer",
                    "savings": 18.00,
                    "reasoning": "Classic LB has zero registered instances. Costs ~$0.025/hr with no use.",
                    "command": f"aws elb delete-load-balancer --load-balancer-name {clb['LoadBalancerName']} --profile {profile} --region {region}",
                    "category": "Networking", "priority": "MEDIUM"
                })

        # --- ElastiCache Clusters ---
        ecache, err = aws_cmd('elasticache describe-cache-clusters --query "CacheClusters[].{CacheClusterId:CacheClusterId,CacheNodeType:CacheNodeType,Engine:Engine,NumCacheNodes:NumCacheNodes,CacheClusterStatus:CacheClusterStatus}"', profile, region)
        if err:
            scan_errors.append(f"ElastiCache/{region}: {err}")
        for cluster in (ecache or []):
            if cluster.get("CacheClusterStatus") != "available":
                continue
            cpu = get_metric(profile, region, "AWS/ElastiCache", "CPUUtilization", "CacheClusterId", cluster["CacheClusterId"], start_date, end_date)
            if cpu and cpu.get("Datapoints"):
                avg = sum(d["Average"] for d in cpu["Datapoints"]) / len(cpu["Datapoints"])
                if avg < 5:
                    findings.append({
                        "id": cluster["CacheClusterId"], "type": f"ElastiCache ({cluster['Engine']})", "region": region,
                        "config": f"{cluster['CacheNodeType']} x{cluster['NumCacheNodes']} (avg CPU: {avg:.1f}%)",
                        "remediation": "Evaluate downsizing or termination - very low utilization",
                        "savings": 25.00,
                        "reasoning": f"ElastiCache cluster averaging {avg:.1f}% CPU. Likely over-provisioned or unused.",
                        "command": f"aws elasticache delete-cache-cluster --cache-cluster-id {cluster['CacheClusterId']} --profile {profile} --region {region}",
                        "category": "Database", "priority": "MEDIUM"
                    })

        # --- OpenSearch Domains ---
        os_domains, err = aws_cmd('opensearch list-domain-names --query "DomainNames[].DomainName"', profile, region)
        if err:
            scan_errors.append(f"OpenSearch/{region}: {err}")
        for domain_name in (os_domains or []):
            cpu = get_metric(profile, region, "AWS/ES", "CPUUtilization", "DomainName", domain_name, start_date, end_date)
            if cpu and cpu.get("Datapoints"):
                avg = sum(d["Average"] for d in cpu["Datapoints"]) / len(cpu["Datapoints"])
                if avg < 5:
                    findings.append({
                        "id": domain_name, "type": "OpenSearch Domain (Underutilized)", "region": region,
                        "config": f"Domain '{domain_name}' (avg CPU: {avg:.1f}%)",
                        "remediation": "Evaluate downsizing or deletion",
                        "savings": 40.00,
                        "reasoning": f"OpenSearch domain averaging {avg:.1f}% CPU over analysis period.",
                        "command": f"aws opensearch delete-domain --domain-name {domain_name} --profile {profile} --region {region}",
                        "category": "Database", "priority": "MEDIUM"
                    })

        # --- Idle Redshift Clusters ---
        redshift, err = aws_cmd('redshift describe-clusters --query "Clusters[].{ClusterIdentifier:ClusterIdentifier,NodeType:NodeType,NumberOfNodes:NumberOfNodes,ClusterStatus:ClusterStatus}"', profile, region)
        if err:
            scan_errors.append(f"Redshift/{region}: {err}")
        for cluster in (redshift or []):
            if cluster.get("ClusterStatus") != "available":
                continue
            conns = get_metric(profile, region, "AWS/Redshift", "DatabaseConnections", "ClusterIdentifier", cluster["ClusterIdentifier"], start_date, end_date)
            if conns and conns.get("Datapoints"):
                avg_conns = sum(d["Average"] for d in conns["Datapoints"]) / len(conns["Datapoints"])
                if avg_conns < 1:
                    findings.append({
                        "id": cluster["ClusterIdentifier"], "type": "Redshift Cluster (Idle)", "region": region,
                        "config": f"{cluster['NodeType']} x{cluster['NumberOfNodes']} (avg connections: {avg_conns:.1f})",
                        "remediation": "Pause or delete idle Redshift cluster",
                        "savings": 180.00,
                        "reasoning": f"Redshift cluster with {avg_conns:.1f} avg connections. Likely unused.",
                        "command": f"aws redshift delete-cluster --cluster-identifier {cluster['ClusterIdentifier']} --skip-final-cluster-snapshot --profile {profile} --region {region}",
                        "category": "Database", "priority": "HIGH"
                    })

        # --- CloudWatch Log Groups with no retention (>1GB) ---
        log_groups, err = aws_cmd('logs describe-log-groups --query "logGroups[?!retentionInDays].{logGroupName:logGroupName,storedBytes:storedBytes}"', profile, region)
        if err:
            scan_errors.append(f"CloudWatch/{region}: {err}")
        for lg in (log_groups or []):
            if (lg.get("storedBytes") or 0) > 1_000_000_000:
                size_gb = lg["storedBytes"] / (1024**3)
                cost = round(size_gb * 0.03, 2)
                findings.append({
                    "id": lg["logGroupName"], "type": "CloudWatch Logs (No Retention)", "region": region,
                    "config": f"{size_gb:.1f}GB stored, no retention policy set",
                    "remediation": "Set retention policy (e.g., 30 days) to reduce storage costs",
                    "savings": cost,
                    "reasoning": "Log group has no retention policy - logs accumulate forever. Set retention to control costs.",
                    "command": f"aws logs put-retention-policy --log-group-name {shlex.quote(lg['logGroupName'])} --retention-in-days 30 --profile {profile} --region {region}",
                    "category": "Storage", "priority": "LOW"
                })

        # --- ECS Services with zero desired count ---
        ecs_clusters, err = aws_cmd('ecs list-clusters --query "clusterArns[]"', profile, region)
        if err:
            scan_errors.append(f"ECS/{region}: {err}")
        for cluster_arn in (ecs_clusters or []):
            services, _ = aws_cmd(f'ecs list-services --cluster {cluster_arn} --query "serviceArns[]"', profile, region)
            if not services:
                continue
            for i in range(0, len(services), 10):
                batch = " ".join(services[i:i+10])
                descs, _ = aws_cmd(f'ecs describe-services --cluster {cluster_arn} --services {batch} --query "services[].{{serviceName:serviceName,desiredCount:desiredCount,runningCount:runningCount,status:status}}"', profile, region)
                for svc in (descs or []):
                    if svc.get("status") == "ACTIVE" and svc.get("desiredCount", 0) == 0 and svc.get("runningCount", 0) == 0:
                        findings.append({
                            "id": svc["serviceName"], "type": "ECS Service (Scaled to Zero)", "region": region,
                            "config": f"Service '{svc['serviceName']}' - 0 desired, 0 running",
                            "remediation": "Delete if no longer needed",
                            "savings": 0.00,
                            "reasoning": "ECS service scaled to zero. No cost currently but clutters the environment.",
                            "command": f"aws ecs delete-service --cluster {cluster_arn} --service {svc['serviceName']} --force --profile {profile} --region {region}",
                            "category": "Compute", "priority": "LOW"
                        })

        # --- EKS Clusters ---
        eks_clusters, err = aws_cmd('eks list-clusters --query "clusters[]"', profile, region)
        if err:
            scan_errors.append(f"EKS/{region}: {err}")
        for cluster_name in (eks_clusters or []):
            nodegroups, _ = aws_cmd(f'eks list-nodegroups --cluster-name {cluster_name} --query "nodegroups[]"', profile, region)
            for ng_name in (nodegroups or []):
                ng, _ = aws_cmd(f'eks describe-nodegroup --cluster-name {cluster_name} --nodegroup-name {ng_name} --query "nodegroup.{{nodegroupName:nodegroupName,instanceTypes:instanceTypes,scalingConfig:scalingConfig,status:status}}"', profile, region)
                if not ng or ng.get("status") != "ACTIVE":
                    continue
                scaling = ng.get("scalingConfig", {})
                desired = scaling.get("desiredSize", 0)
                min_size = scaling.get("minSize", 0)
                max_size = scaling.get("maxSize", 0)
                instance_types = ng.get("instanceTypes", [])
                if desired > 0 and min_size == max_size:
                    findings.append({
                        "id": f"{cluster_name}/{ng_name}", "type": "EKS Nodegroup (Fixed Size)", "region": region,
                        "config": f"{','.join(instance_types)} min={min_size} max={max_size} desired={desired}",
                        "remediation": "Enable autoscaling or evaluate if over-provisioned",
                        "savings": 0.00,
                        "reasoning": f"Nodegroup has fixed size (min=max={min_size}). No autoscaling configured.",
                        "command": f"aws eks update-nodegroup-config --cluster-name {cluster_name} --nodegroup-name {ng_name} --scaling-config minSize=1,maxSize={max_size},desiredSize={min_size} --profile {profile} --region {region}",
                        "category": "Compute", "priority": "LOW"
                    })
                if desired == 0:
                    findings.append({
                        "id": f"{cluster_name}/{ng_name}", "type": "EKS Nodegroup (Scaled to Zero)", "region": region,
                        "config": f"{','.join(instance_types)} - scaled to 0 nodes",
                        "remediation": "Delete if no longer needed",
                        "savings": 0.00,
                        "reasoning": "EKS nodegroup scaled to zero. Cluster control plane still billed ($73/mo).",
                        "command": f"aws eks delete-nodegroup --cluster-name {cluster_name} --nodegroup-name {ng_name} --profile {profile} --region {region}",
                        "category": "Compute", "priority": "LOW"
                    })
            if not (nodegroups or []):
                findings.append({
                    "id": cluster_name, "type": "EKS Cluster (No Nodegroups)", "region": region,
                    "config": f"EKS cluster '{cluster_name}' with no nodegroups",
                    "remediation": "Delete empty EKS cluster to save control plane costs",
                    "savings": 73.00,
                    "reasoning": "EKS cluster running with no nodegroups. Control plane costs $0.10/hr ($73/mo).",
                    "command": f"aws eks delete-cluster --name {cluster_name} --profile {profile} --region {region}",
                    "category": "Compute", "priority": "HIGH"
                })

        # --- SageMaker Endpoints ---
        sm_endpoints, err = aws_cmd('sagemaker list-endpoints --query "Endpoints[].{EndpointName:EndpointName,EndpointStatus:EndpointStatus}"', profile, region)
        if err:
            scan_errors.append(f"SageMaker/{region}: {err}")
        for ep in (sm_endpoints or []):
            if ep.get("EndpointStatus") != "InService":
                continue
            invocations = get_metric(profile, region, "AWS/SageMaker", "Invocations", "EndpointName", ep["EndpointName"], start_date, end_date)
            if invocations:
                avg_inv = 0
                if invocations.get("Datapoints"):
                    avg_inv = sum(d["Average"] for d in invocations["Datapoints"]) / len(invocations["Datapoints"])
                if avg_inv < 1:
                    findings.append({
                        "id": ep["EndpointName"], "type": "SageMaker Endpoint (Idle)", "region": region,
                        "config": f"Endpoint '{ep['EndpointName']}' - {'near-zero' if avg_inv > 0 else 'zero'} invocations",
                        "remediation": "Delete idle SageMaker endpoint",
                        "savings": 50.00,
                        "reasoning": f"SageMaker endpoint averaging {avg_inv:.2f} invocations. Billed for instance hours with no traffic.",
                        "command": f"aws sagemaker delete-endpoint --endpoint-name {ep['EndpointName']} --profile {profile} --region {region}",
                        "category": "AI/ML", "priority": "HIGH"
                    })

        # --- SageMaker Notebook Instances ---
        sm_notebooks, err = aws_cmd('sagemaker list-notebook-instances --query "NotebookInstances[].{NotebookInstanceName:NotebookInstanceName,NotebookInstanceStatus:NotebookInstanceStatus,InstanceType:InstanceType}"', profile, region)
        if err:
            scan_errors.append(f"SageMaker-NB/{region}: {err}")
        for nb in (sm_notebooks or []):
            if nb.get("NotebookInstanceStatus") == "InService":
                findings.append({
                    "id": nb["NotebookInstanceName"], "type": "SageMaker Notebook (Running)", "region": region,
                    "config": f"{nb['InstanceType']} - running",
                    "remediation": "Stop notebook instance when not in use",
                    "savings": 30.00,
                    "reasoning": "SageMaker notebook running continuously. Stop when not actively used.",
                    "command": f"aws sagemaker stop-notebook-instance --notebook-instance-name {nb['NotebookInstanceName']} --profile {profile} --region {region}",
                    "category": "AI/ML", "priority": "MEDIUM"
                })
            elif nb.get("NotebookInstanceStatus") == "Stopped":
                findings.append({
                    "id": nb["NotebookInstanceName"], "type": "SageMaker Notebook (Stopped)", "region": region,
                    "config": f"{nb['InstanceType']} - stopped (storage still billed)",
                    "remediation": "Delete if no longer needed to eliminate storage charges",
                    "savings": 5.00,
                    "reasoning": "Stopped notebook still incurs EBS storage charges. Delete if not needed.",
                    "command": f"aws sagemaker delete-notebook-instance --notebook-instance-name {nb['NotebookInstanceName']} --profile {profile} --region {region}",
                    "category": "AI/ML", "priority": "LOW"
                })

        # --- Bedrock Provisioned Model Throughput ---
        bedrock_pts, err = aws_cmd('bedrock list-provisioned-model-throughputs --query "provisionedModelSummaries[].{provisionedModelName:provisionedModelName,provisionedModelArn:provisionedModelArn,status:status,modelUnits:modelUnits}"', profile, region)
        if err:
            scan_errors.append(f"Bedrock/{region}: {err}")
        for pt in (bedrock_pts or []):
            if pt.get("status") == "InService":
                units = pt.get("modelUnits", 1)
                findings.append({
                    "id": pt["provisionedModelName"], "type": "Bedrock Provisioned Throughput", "region": region,
                    "config": f"{units} model unit(s) provisioned",
                    "remediation": "Delete if not actively needed - very expensive",
                    "savings": round(units * 25 * 730, 2),
                    "reasoning": f"Bedrock provisioned throughput with {units} unit(s). Costs ~$25/hr/unit.",
                    "command": f"aws bedrock delete-provisioned-model-throughput --provisioned-model-id {pt['provisionedModelArn']} --profile {profile} --region {region}",
                    "category": "AI/ML", "priority": "HIGH"
                })

        # --- SageMaker Training Jobs (stuck >24h) ---
        training_jobs, _ = aws_cmd('sagemaker list-training-jobs --status-equals InProgress --query "TrainingJobSummaries[].{TrainingJobName:TrainingJobName,CreationTime:CreationTime}"', profile, region)
        for tj in (training_jobs or []):
            created = tj.get("CreationTime", "")
            if created:
                try:
                    created_dt = datetime.fromisoformat(created.replace("Z", "+00:00").replace("+00:00", ""))
                    hours_running = (datetime.utcnow() - created_dt).total_seconds() / 3600
                    if hours_running > 24:
                        findings.append({
                            "id": tj["TrainingJobName"], "type": "SageMaker Training Job (Long-Running)", "region": region,
                            "config": f"Running for {hours_running:.0f} hours",
                            "remediation": "Investigate - may be stuck",
                            "savings": 20.00,
                            "reasoning": f"Training job running for {hours_running:.0f}+ hours. May be stuck.",
                            "command": f"aws sagemaker stop-training-job --training-job-name {tj['TrainingJobName']} --profile {profile} --region {region}",
                            "category": "AI/ML", "priority": "HIGH"
                        })
                except (ValueError, TypeError):
                    pass

    # --- Build Excel Report ---
    total_savings = sum(f["savings"] for f in findings)
    target_spend = projected_monthly - total_savings
    reduction_pct = (total_savings / projected_monthly * 100) if projected_monthly > 0 else 0

    categories = {}
    for f in findings:
        cat = f["category"]
        if cat not in categories:
            categories[cat] = {"waste": 0, "savings": 0, "priority": f["priority"]}
        categories[cat]["waste"] += f["savings"]
        categories[cat]["savings"] += f["savings"]

    risk_map = {
        "Database": "Medium (Requires App Validation)",
        "Networking": "Medium (Connectivity Impact)",
        "Compute": "Low (Underutilized)",
        "Storage": "Low (No Downtime)",
        "Migration": "Low (No Active Tasks)",
        "AI/ML": "Medium (Model Serving Impact)",
    }

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1F3864")
    money_fmt = '$#,##0.00'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    priority_fills = {
        "HIGH": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "LOW": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    }

    def style_header_row(ws, row, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
            ws.column_dimensions[col_letter].width = max_len + 3

    output_dir = Path.home() / "finops-reports"
    output_dir.mkdir(exist_ok=True)
    wb = Workbook()

    # --- Sheet 1: Executive Summary ---
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = "1F4E79"
    ws1["A1"] = "AWS FinOps — Cost Optimization Report"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:F1")
    ws1["A3"] = "AWS Account ID"
    ws1["B3"] = "Report Date"
    ws1["C3"] = "Monthly Spend (Projected)"
    ws1["D3"] = "Target Monthly Spend"
    ws1["E3"] = "Total Potential Savings"
    ws1["F3"] = "Reduction %"
    style_header_row(ws1, 3, 6)
    ws1["A4"] = account_id
    ws1["B4"] = today
    ws1["C4"] = projected_monthly
    ws1["C4"].number_format = money_fmt
    ws1["D4"] = target_spend
    ws1["D4"].number_format = money_fmt
    ws1["E4"] = total_savings
    ws1["E4"].number_format = money_fmt
    ws1["F4"] = f"{reduction_pct:.1f}%"

    ws1["A6"] = "Category Breakdown"
    ws1["A6"].font = Font(bold=True, size=12)
    for i, h in enumerate(["Resource Category", "Monthly Waste ($)", "Potential Savings ($)", "Priority", "Risk Impact"], 1):
        ws1.cell(row=7, column=i, value=h)
    style_header_row(ws1, 7, 5)
    row = 8
    for cat, data in categories.items():
        ws1.cell(row=row, column=1, value=cat)
        ws1.cell(row=row, column=2, value=data["waste"]).number_format = money_fmt
        ws1.cell(row=row, column=3, value=data["savings"]).number_format = money_fmt
        cell = ws1.cell(row=row, column=4, value=data["priority"])
        cell.fill = priority_fills.get(data["priority"], PatternFill())
        ws1.cell(row=row, column=5, value=risk_map.get(cat, "Low"))
        for col in range(1, 6):
            ws1.cell(row=row, column=col).border = thin_border
        row += 1

    # Scan coverage info
    row += 1
    ws1.cell(row=row, column=1, value="Scan Coverage").font = Font(bold=True, size=12)
    row += 1
    ws1.cell(row=row, column=1, value=f"Regions scanned: {', '.join(regions)}")
    row += 1
    ws1.cell(row=row, column=1, value=f"Resource types checked: 20")
    row += 1
    ws1.cell(row=row, column=1, value=f"API errors encountered: {len(scan_errors)}")
    if scan_errors:
        row += 1
        ws1.cell(row=row, column=1, value="Failed checks:").font = Font(bold=True)
        for err_msg in scan_errors[:10]:
            row += 1
            ws1.cell(row=row, column=1, value=f"  ⚠ {err_msg}")
    auto_width(ws1)

    # --- Sheet 2: Month-to-Date Costs ---
    ws_mtd = wb.create_sheet("Cost - Month-to-Date")
    ws_mtd.sheet_properties.tabColor = "2E75B6"
    ws_mtd["A1"] = f"AWS Cost Breakdown — Month-to-Date ({mtd_start} to {today})"
    ws_mtd["A1"].font = title_font
    ws_mtd.merge_cells("A1:D1")
    ws_mtd["A2"] = f"Account: {account_id}"
    ws_mtd["A2"].font = Font(italic=True, size=10)
    for i, h in enumerate(["Service", "MTD Cost ($)", "% of Total", "Daily Average ($)"], 1):
        ws_mtd.cell(row=4, column=i, value=h)
    style_header_row(ws_mtd, 4, 4)
    row = 5
    for svc, cost in sorted(mtd_by_service.items(), key=lambda x: x[1], reverse=True):
        if cost < 0.01:
            continue
        ws_mtd.cell(row=row, column=1, value=svc)
        ws_mtd.cell(row=row, column=2, value=round(cost, 2)).number_format = money_fmt
        pct = (cost / mtd_total * 100) if mtd_total > 0 else 0
        ws_mtd.cell(row=row, column=3, value=f"{pct:.1f}%")
        ws_mtd.cell(row=row, column=4, value=round(cost / days_in_month_so_far, 2)).number_format = money_fmt
        for col in range(1, 5):
            ws_mtd.cell(row=row, column=col).border = thin_border
        row += 1
    ws_mtd.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws_mtd.cell(row=row, column=2, value=round(mtd_total, 2)).number_format = money_fmt
    ws_mtd.cell(row=row, column=2).font = Font(bold=True)
    ws_mtd.cell(row=row, column=3, value="100%").font = Font(bold=True)
    ws_mtd.cell(row=row, column=4, value=round(mtd_total / days_in_month_so_far, 2)).number_format = money_fmt
    ws_mtd.cell(row=row, column=4).font = Font(bold=True)
    for col in range(1, 5):
        ws_mtd.cell(row=row, column=col).border = thin_border
    auto_width(ws_mtd)

    # --- Sheet 3: Last 3 Months ---
    ws_l3m = wb.create_sheet("Cost - Last 3 Months")
    ws_l3m.sheet_properties.tabColor = "70AD47"
    ws_l3m["A1"] = f"AWS Cost Breakdown — Last 3 Months ({three_months_ago} to {mtd_start})"
    ws_l3m["A1"].font = title_font
    col_count_l3m = 3 + len(l3m_months)
    if col_count_l3m > 1:
        ws_l3m.merge_cells(f"A1:{get_column_letter(col_count_l3m)}1")
    ws_l3m["A2"] = f"Account: {account_id}"
    ws_l3m["A2"].font = Font(italic=True, size=10)
    headers_l3m = ["Service"] + l3m_months + ["Total ($)", "Avg Monthly ($)"]
    for i, h in enumerate(headers_l3m, 1):
        ws_l3m.cell(row=4, column=i, value=h)
    style_header_row(ws_l3m, 4, len(headers_l3m))
    svc_totals = {svc: sum(months.values()) for svc, months in l3m_by_service_month.items()}
    row = 5
    grand_totals = {m: 0.0 for m in l3m_months}
    grand_total = 0.0
    for svc, total in sorted(svc_totals.items(), key=lambda x: x[1], reverse=True):
        if total < 0.01:
            continue
        ws_l3m.cell(row=row, column=1, value=svc)
        for mi, month in enumerate(l3m_months, 2):
            amt = l3m_by_service_month[svc].get(month, 0)
            ws_l3m.cell(row=row, column=mi, value=round(amt, 2)).number_format = money_fmt
            grand_totals[month] += amt
        ws_l3m.cell(row=row, column=len(l3m_months) + 2, value=round(total, 2)).number_format = money_fmt
        ws_l3m.cell(row=row, column=len(l3m_months) + 3, value=round(total / max(len(l3m_months), 1), 2)).number_format = money_fmt
        grand_total += total
        for col in range(1, len(headers_l3m) + 1):
            ws_l3m.cell(row=row, column=col).border = thin_border
        row += 1
    ws_l3m.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for mi, month in enumerate(l3m_months, 2):
        ws_l3m.cell(row=row, column=mi, value=round(grand_totals[month], 2)).number_format = money_fmt
        ws_l3m.cell(row=row, column=mi).font = Font(bold=True)
    ws_l3m.cell(row=row, column=len(l3m_months) + 2, value=round(grand_total, 2)).number_format = money_fmt
    ws_l3m.cell(row=row, column=len(l3m_months) + 2).font = Font(bold=True)
    ws_l3m.cell(row=row, column=len(l3m_months) + 3, value=round(grand_total / max(len(l3m_months), 1), 2)).number_format = money_fmt
    ws_l3m.cell(row=row, column=len(l3m_months) + 3).font = Font(bold=True)
    for col in range(1, len(headers_l3m) + 1):
        ws_l3m.cell(row=row, column=col).border = thin_border
    auto_width(ws_l3m)

    # --- Sheet 4: Remediation Ledger ---
    ws2 = wb.create_sheet("Remediation Ledger")
    ws2.sheet_properties.tabColor = "C00000"
    ws2["A1"] = "Technical Remediation Ledger"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:H1")
    headers2 = ["Resource ID", "Resource Type", "Region", "Current Configuration",
                "Suggested Remediation", "Est. Monthly Savings ($)", "Reasoning", "Remediation Command"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=i, value=h)
    style_header_row(ws2, 3, 8)
    for idx, f in enumerate(findings, 4):
        ws2.cell(row=idx, column=1, value=f["id"])
        ws2.cell(row=idx, column=2, value=f["type"])
        ws2.cell(row=idx, column=3, value=f["region"])
        ws2.cell(row=idx, column=4, value=f["config"])
        ws2.cell(row=idx, column=5, value=f["remediation"])
        ws2.cell(row=idx, column=6, value=f["savings"]).number_format = money_fmt
        ws2.cell(row=idx, column=7, value=f["reasoning"])
        ws2.cell(row=idx, column=8, value=f["command"])
        for col in range(1, 9):
            ws2.cell(row=idx, column=col).border = thin_border
            ws2.cell(row=idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    auto_width(ws2)

    # Save
    filename = f"finops-report-{account_id}-{today}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)

    # Save history for trend tracking
    save_history(account_id, today, findings, mtd_total, projected_monthly, total_savings)

    logger.info(f"Analysis complete: {len(findings)} findings, ${total_savings:.2f} savings, {len(scan_errors)} errors")

    # --- Text Summary ---
    summary = f"✅ Report saved to: ~/finops-reports/{filename}\n"
    summary += f"   Sheets: Executive Summary | Cost - Month-to-Date | Cost - Last 3 Months | Remediation Ledger\n\n"
    summary += f"Account: {account_id} | Date: {today}\n"
    summary += f"MTD Spend: ${mtd_total:.2f} | Projected Monthly: ${projected_monthly:.2f}\n"
    summary += f"Total Potential Savings: ${total_savings:.2f} ({reduction_pct:.1f}% reduction)\n"
    summary += f"Scan Coverage: {len(regions)} region(s), 20 resource types, {len(scan_errors)} API errors\n\n"

    if scan_errors:
        summary += f"⚠ Failed checks ({len(scan_errors)}):\n"
        for err_msg in scan_errors[:5]:
            summary += f"  - {err_msg}\n"
        if len(scan_errors) > 5:
            summary += f"  ... and {len(scan_errors) - 5} more (see log file)\n"
        summary += "\n"

    summary += "Findings Summary:\n"
    for f in findings:
        summary += f"  • [{f['priority']}] {f['type']} - {f['id']} → ${f['savings']:.2f}/mo\n"
        summary += f"    Remediation: {f['remediation']}\n"
        summary += f"    Command: {f['command']}\n\n"

    if not findings:
        summary += "  No cost optimization opportunities found.\n"

    return summary


@mcp.tool()
def analyze_org_costs(regions: list[str], role_name: str = "OrganizationAccountAccessRole", lookback_days: int = 30) -> str:
    """
    Analyze ALL accounts in an AWS Organization for cost optimization.
    Uses the management account credentials from environment variables
    (AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, AWS_SESSION_TOKEN) and
    assumes a role into each member account to scan resources.

    Produces a single Excel report with separate sheet tabs per account.

    Args:
        regions: List of AWS regions to scan (e.g. ['ap-south-1'])
        role_name: IAM role name to assume in member accounts (default: OrganizationAccountAccessRole)
        lookback_days: Number of days to look back for utilization metrics (default: 30)
    """
    import os

    region_pattern = re.compile(r"^[a-z]{2}-[a-z]+-\d+$")
    for r in regions:
        if not region_pattern.match(r):
            return f"❌ Invalid region format: '{r}'. Expected format: us-east-1, ap-south-1, etc."

    # Verify management account credentials from env
    identity, err = aws_cmd("sts get-caller-identity", "", regions[0])
    if err:
        return f"❌ Cannot authenticate. Export AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY (and optionally AWS_SESSION_TOKEN) for the management account.\nError: {err}"
    mgmt_account_id = identity.get("Account", "unknown")
    logger.info(f"Org scan: management account={mgmt_account_id}, regions={regions}")

    # List all accounts in the organization
    accounts_data, err = aws_cmd("organizations list-accounts --query \"Accounts[?Status=='ACTIVE'].[Id,Name]\"", "", regions[0])
    if err:
        return f"❌ Cannot list organization accounts. Ensure this is the management account with Organizations access.\nError: {err}"

    if not accounts_data:
        return "❌ No accounts found in the organization."

    accounts = [{"id": a[0], "name": a[1]} for a in accounts_data]
    logger.info(f"Found {len(accounts)} accounts in organization")

    end_date = datetime.utcnow().strftime("%Y-%m-%d")
    start_date = (datetime.utcnow() - timedelta(days=lookback_days)).strftime("%Y-%m-%d")
    today = end_date
    mtd_start = datetime.utcnow().strftime("%Y-%m-01")
    three_months_ago = (datetime.utcnow().replace(day=1) - timedelta(days=90)).strftime("%Y-%m-01")
    days_in_month_so_far = max((datetime.utcnow() - datetime.strptime(mtd_start, "%Y-%m-%d")).days, 1)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1F3864")
    money_fmt = '$#,##0.00'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    priority_fills = {
        "HIGH": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "LOW": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    }
    tab_colors = ["1F4E79", "2E75B6", "70AD47", "C00000", "7030A0", "ED7D31", "44546A", "4472C4"]

    def style_header_row(ws, row, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
            ws.column_dimensions[col_letter].width = max_len + 3

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    all_findings = []
    account_summaries = []
    scan_errors = []

    for idx, account in enumerate(accounts):
        acct_id = account["id"]
        acct_name = account["name"]
        role_arn = f"arn:aws:iam::{acct_id}:role/{role_name}"

        logger.info(f"Scanning account: {acct_name} ({acct_id})")

        # For management account, use env creds directly; for others, assume role
        if acct_id == mgmt_account_id:
            cmd_fn = lambda cmd, region: aws_cmd(cmd, "", region)
        else:
            cmd_fn = lambda cmd, region, ra=role_arn: aws_cmd_with_role(cmd, ra, region)

        # Quick connectivity test
        test_identity, err = cmd_fn("sts get-caller-identity", regions[0])
        if err:
            scan_errors.append(f"{acct_name} ({acct_id}): Cannot assume role - {err}")
            logger.warning(f"Skipping {acct_name}: {err}")
            continue

        # Scan resources in this account
        findings = []
        for region in regions:
            # EC2
            instances, err = cmd_fn(
                'ec2 describe-instances --query "Reservations[].Instances[].{InstanceId:InstanceId,InstanceType:InstanceType,State:State.Name}"',
                region
            )
            if err:
                scan_errors.append(f"{acct_name}/EC2/{region}: {err}")
            for inst in (instances or []):
                if inst["State"] == "stopped":
                    findings.append({
                        "id": inst["InstanceId"], "type": "EC2 (Stopped)", "region": region,
                        "config": f"{inst['InstanceType']} (stopped)",
                        "remediation": "Terminate if not needed",
                        "savings": 0.80, "category": "Compute", "priority": "MEDIUM"
                    })

            # EBS
            volumes, err = cmd_fn(
                'ec2 describe-volumes --query "Volumes[?State==\'available\'].{VolumeId:VolumeId,Size:Size,VolumeType:VolumeType}"',
                region
            )
            if err:
                scan_errors.append(f"{acct_name}/EBS/{region}: {err}")
            for vol in (volumes or []):
                cost = get_ebs_monthly_cost(vol["Size"], vol["VolumeType"], region)
                findings.append({
                    "id": vol["VolumeId"], "type": "EBS (Unattached)", "region": region,
                    "config": f"{vol['Size']}GB {vol['VolumeType']}",
                    "remediation": "Delete unattached volume",
                    "savings": cost, "category": "Storage", "priority": "MEDIUM"
                })

            # RDS
            rds_instances, err = cmd_fn(
                'rds describe-db-instances --query "DBInstances[].{DBInstanceIdentifier:DBInstanceIdentifier,DBInstanceClass:DBInstanceClass,Engine:Engine}"',
                region
            )
            if err:
                scan_errors.append(f"{acct_name}/RDS/{region}: {err}")
            for db in (rds_instances or []):
                findings.append({
                    "id": db["DBInstanceIdentifier"], "type": f"RDS ({db['Engine']})", "region": region,
                    "config": f"{db['DBInstanceClass']}",
                    "remediation": "Check utilization",
                    "savings": 0, "category": "Database", "priority": "LOW"
                })

        # Create sheet for this account
        sheet_name = f"{acct_name[:20]} ({acct_id[-4:]})"
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab_colors[idx % len(tab_colors)]

        ws["A1"] = f"Account: {acct_name} ({acct_id})"
        ws["A1"].font = title_font
        ws.merge_cells("A1:F1")

        # Findings table
        headers = ["Resource ID", "Type", "Region", "Configuration", "Suggested Remediation", "Est. Savings ($)"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        style_header_row(ws, 3, 6)

        acct_savings = 0
        for fi, f in enumerate(findings, 4):
            ws.cell(row=fi, column=1, value=f["id"])
            ws.cell(row=fi, column=2, value=f["type"])
            ws.cell(row=fi, column=3, value=f["region"])
            ws.cell(row=fi, column=4, value=f["config"])
            ws.cell(row=fi, column=5, value=f["remediation"])
            ws.cell(row=fi, column=6, value=f["savings"]).number_format = money_fmt
            acct_savings += f["savings"]
            for col in range(1, 7):
                ws.cell(row=fi, column=col).border = thin_border

        if not findings:
            ws.cell(row=4, column=1, value="No optimization opportunities found.")

        auto_width(ws)
        all_findings.extend(findings)
        account_summaries.append({
            "name": acct_name, "id": acct_id,
            "findings": len(findings), "savings": acct_savings
        })

    # Add Org Summary sheet at the beginning
    ws_summary = wb.create_sheet("Org Summary", 0)
    ws_summary.sheet_properties.tabColor = "000000"
    ws_summary["A1"] = "AWS Organization — Cost Optimization Summary"
    ws_summary["A1"].font = title_font
    ws_summary.merge_cells("A1:E1")
    ws_summary["A2"] = f"Management Account: {mgmt_account_id} | Date: {today} | Regions: {', '.join(regions)}"
    ws_summary["A2"].font = Font(italic=True, size=10)

    headers = ["Account Name", "Account ID", "Findings", "Est. Monthly Savings ($)", "Status"]
    for i, h in enumerate(headers, 1):
        ws_summary.cell(row=4, column=i, value=h)
    style_header_row(ws_summary, 4, 5)

    row = 5
    total_org_savings = 0
    for acct in account_summaries:
        ws_summary.cell(row=row, column=1, value=acct["name"])
        ws_summary.cell(row=row, column=2, value=acct["id"])
        ws_summary.cell(row=row, column=3, value=acct["findings"])
        ws_summary.cell(row=row, column=4, value=acct["savings"]).number_format = money_fmt
        ws_summary.cell(row=row, column=5, value="Scanned ✓")
        total_org_savings += acct["savings"]
        for col in range(1, 6):
            ws_summary.cell(row=row, column=col).border = thin_border
        row += 1

    # Errors
    for err_msg in scan_errors:
        if "Cannot assume role" in err_msg:
            acct_info = err_msg.split(":")[0]
            ws_summary.cell(row=row, column=1, value=acct_info)
            ws_summary.cell(row=row, column=5, value="⚠ Access Denied")
            ws_summary.cell(row=row, column=5).font = Font(color="FF0000")
            for col in range(1, 6):
                ws_summary.cell(row=row, column=col).border = thin_border
            row += 1

    # Total row
    ws_summary.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=row, column=3, value=len(all_findings)).font = Font(bold=True)
    ws_summary.cell(row=row, column=4, value=total_org_savings).number_format = money_fmt
    ws_summary.cell(row=row, column=4).font = Font(bold=True)
    for col in range(1, 6):
        ws_summary.cell(row=row, column=col).border = thin_border
    auto_width(ws_summary)

    # Save
    output_dir = Path.home() / "finops-reports"
    output_dir.mkdir(exist_ok=True)
    filename = f"finops-org-report-{mgmt_account_id}-{today}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)

    summary = f"✅ Organization report saved to: ~/finops-reports/{filename}\n\n"
    summary += f"Organization: {len(accounts)} accounts discovered\n"
    summary += f"Successfully scanned: {len(account_summaries)} accounts\n"
    summary += f"Failed: {len(scan_errors)} errors\n"
    summary += f"Total findings: {len(all_findings)}\n"
    summary += f"Total potential savings: ${total_org_savings:.2f}/mo\n\n"
    summary += "Per-account breakdown:\n"
    for acct in account_summaries:
        summary += f"  • {acct['name']} ({acct['id']}): {acct['findings']} findings, ${acct['savings']:.2f}/mo\n"

    if scan_errors:
        summary += f"\n⚠ Errors ({len(scan_errors)}):\n"
        for err_msg in scan_errors[:5]:
            summary += f"  - {err_msg}\n"

    return summary


@mcp.tool()
def execute_remediation(command: str, dry_run: bool = True) -> str:
    """
    Execute a specific AWS CLI remediation command from the analysis output.
    Commands are validated against a whitelist of safe operations.

    Args:
        command: The full AWS CLI command to execute (from the remediation report)
        dry_run: If True (default), only validates and shows what would run. Set to False to execute.
    """
    # Validate against whitelist
    if not validate_command(command):
        return (
            f"❌ Command rejected - not in allowed remediation whitelist.\n"
            f"Command: {command}\n\n"
            f"Only these operations are allowed:\n"
            f"  • EC2: terminate, stop, delete-volume, release-address, delete-nat-gateway, delete-vpn, delete-snapshot\n"
            f"  • RDS: delete-db-instance\n"
            f"  • ELB: delete-load-balancer\n"
            f"  • ElastiCache: delete-cache-cluster\n"
            f"  • OpenSearch: delete-domain\n"
            f"  • Redshift: delete-cluster\n"
            f"  • DMS: delete-replication-instance\n"
            f"  • CloudWatch: put-retention-policy\n"
            f"  • ECS: delete-service\n"
            f"  • EKS: delete-cluster, delete-nodegroup, update-nodegroup-config\n"
            f"  • SageMaker: delete-endpoint, stop/delete-notebook, stop-training-job\n"
            f"  • Bedrock: delete-provisioned-model-throughput\n"
        )

    if dry_run:
        return (
            f"🔍 DRY RUN — Command validated successfully:\n"
            f"  {command}\n\n"
            f"To execute for real, call again with dry_run=False"
        )

    logger.info(f"Executing remediation: {command}")
    try:
        args = shlex.split(command)
        result = subprocess.run(args, capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            logger.info(f"Remediation succeeded: {command}")
            return f"✅ Command executed successfully:\n{result.stdout}"
        else:
            logger.warning(f"Remediation failed: {command} -> {result.stderr}")
            return f"❌ Command failed (exit {result.returncode}):\n{result.stderr}"
    except subprocess.TimeoutExpired:
        return "❌ Command timed out after 60 seconds"
    except Exception as e:
        return f"❌ Error: {str(e)}"


@mcp.tool()
def get_savings_history(profile: str) -> str:
    """
    Show historical trend of cost optimization findings across runs.
    Useful for tracking progress over time.

    Args:
        profile: AWS CLI profile name to get account ID
    """
    # Get account ID
    identity, err = aws_cmd("sts get-caller-identity", profile, "us-east-1")
    if err:
        return f"❌ Cannot authenticate: {err}"
    account_id = identity.get("Account", "unknown")

    history_file = LOG_DIR / f"history-{account_id}.json"
    if not history_file.exists():
        return f"No history found for account {account_id}. Run analyze_aws_costs first."

    history = json.loads(history_file.read_text())
    if not history:
        return "History file is empty."

    summary = f"📊 Savings History for Account {account_id}\n"
    summary += f"{'='*60}\n\n"
    summary += f"{'Date':<12} {'MTD Spend':>10} {'Projected':>10} {'Savings':>10} {'Findings':>8}\n"
    summary += f"{'-'*12} {'-'*10} {'-'*10} {'-'*10} {'-'*8}\n"

    for entry in history[-10:]:  # Last 10 runs
        summary += (
            f"{entry['date']:<12} "
            f"${entry['mtd_spend']:>8.2f} "
            f"${entry['projected_monthly']:>8.2f} "
            f"${entry['total_savings_identified']:>8.2f} "
            f"{entry['finding_count']:>8}\n"
        )

    if len(history) >= 2:
        first = history[0]
        last = history[-1]
        trend = last["total_savings_identified"] - first["total_savings_identified"]
        summary += f"\nTrend: Savings identified {'increased' if trend > 0 else 'decreased'} by ${abs(trend):.2f} since first scan.\n"

    return summary


def main():
    mcp.run()


if __name__ == "__main__":
    main()
